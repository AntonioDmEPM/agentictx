"""Business Case module service — all DB operations and Excel generation."""
import io
import uuid
from typing import Any

from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from sqlalchemy.orm.attributes import flag_modified

from app.agents.business_case_agent import compute_scenario
from app.models.business_case import BusinessCase, BusinessCaseScenario
from app.schemas.business_case import (
    AssumptionsUpdate,
    BusinessCaseRead,
    BusinessCaseScenarioCreate,
    BusinessCaseScenarioRead,
    BusinessCaseScenarioUpdate,
    ModalityUpdate,
    ScenarioResults,
)


# ─── Internal helpers ─────────────────────────────────────────────────────────

async def _get_bc_with_scenarios(
    db: AsyncSession, use_case_id: uuid.UUID
) -> BusinessCase | None:
    result = await db.execute(
        select(BusinessCase)
        .where(BusinessCase.use_case_id == use_case_id)
        .options(selectinload(BusinessCase.scenarios))
    )
    return result.scalar_one_or_none()


# ─── Get or create ────────────────────────────────────────────────────────────

async def get_or_create_business_case(
    db: AsyncSession, use_case_id: uuid.UUID
) -> BusinessCaseRead:
    bc = await _get_bc_with_scenarios(db, use_case_id)
    if bc is None:
        bc = BusinessCase(use_case_id=use_case_id, coverage_ramp=[])
        db.add(bc)
        await db.flush()
        await db.refresh(bc)
        # Re-fetch with scenarios loaded (empty list after create)
        bc = await _get_bc_with_scenarios(db, use_case_id)
    return BusinessCaseRead.model_validate(bc)


# ─── Assumptions update ───────────────────────────────────────────────────────

async def update_assumptions(
    db: AsyncSession,
    use_case_id: uuid.UUID,
    payload: AssumptionsUpdate,
) -> BusinessCaseRead | None:
    bc = await _get_bc_with_scenarios(db, use_case_id)
    if bc is None:
        return None
    for field, value in payload.model_dump(exclude_none=True).items():
        setattr(bc, field, value)
    await db.flush()
    await db.refresh(bc)
    bc = await _get_bc_with_scenarios(db, use_case_id)
    return BusinessCaseRead.model_validate(bc)


# ─── Modality update ──────────────────────────────────────────────────────────

async def update_modality(
    db: AsyncSession,
    use_case_id: uuid.UUID,
    payload: ModalityUpdate,
) -> BusinessCaseRead | None:
    bc = await _get_bc_with_scenarios(db, use_case_id)
    if bc is None:
        return None
    for field, value in payload.model_dump(exclude_none=True).items():
        setattr(bc, field, value)
    await db.flush()
    await db.refresh(bc)
    bc = await _get_bc_with_scenarios(db, use_case_id)
    return BusinessCaseRead.model_validate(bc)


# ─── Scenario CRUD ────────────────────────────────────────────────────────────

async def create_scenario(
    db: AsyncSession,
    use_case_id: uuid.UUID,
    payload: BusinessCaseScenarioCreate,
) -> BusinessCaseScenarioRead | None:
    bc = await _get_bc_with_scenarios(db, use_case_id)
    if bc is None:
        return None

    # Auto-increment sort_order
    max_order_result = await db.execute(
        select(func.max(BusinessCaseScenario.sort_order)).where(
            BusinessCaseScenario.business_case_id == bc.id
        )
    )
    current_max = max_order_result.scalar() or -1
    next_order = current_max + 1

    scenario = BusinessCaseScenario(
        business_case_id=bc.id,
        sort_order=next_order,
        results={},
        **payload.model_dump(),
    )
    db.add(scenario)
    await db.flush()
    await db.refresh(scenario)
    return BusinessCaseScenarioRead.model_validate(scenario)


async def update_scenario(
    db: AsyncSession,
    use_case_id: uuid.UUID,
    scenario_id: uuid.UUID,
    payload: BusinessCaseScenarioUpdate,
) -> BusinessCaseScenarioRead | None:
    bc = await _get_bc_with_scenarios(db, use_case_id)
    if bc is None:
        return None

    result = await db.execute(
        select(BusinessCaseScenario).where(
            BusinessCaseScenario.id == scenario_id,
            BusinessCaseScenario.business_case_id == bc.id,
        )
    )
    scenario = result.scalar_one_or_none()
    if scenario is None:
        return None

    for field, value in payload.model_dump(exclude_none=True).items():
        setattr(scenario, field, value)
    await db.flush()
    await db.refresh(scenario)
    return BusinessCaseScenarioRead.model_validate(scenario)


async def delete_scenario(
    db: AsyncSession,
    use_case_id: uuid.UUID,
    scenario_id: uuid.UUID,
) -> bool:
    bc = await _get_bc_with_scenarios(db, use_case_id)
    if bc is None:
        return False

    result = await db.execute(
        select(BusinessCaseScenario).where(
            BusinessCaseScenario.id == scenario_id,
            BusinessCaseScenario.business_case_id == bc.id,
        )
    )
    scenario = result.scalar_one_or_none()
    if scenario is None:
        return False
    await db.delete(scenario)
    return True


# ─── Calculate ────────────────────────────────────────────────────────────────

async def calculate_business_case(
    db: AsyncSession,
    use_case_id: uuid.UUID,
) -> BusinessCaseRead | None:
    bc = await _get_bc_with_scenarios(db, use_case_id)
    if bc is None:
        return None

    for scenario in bc.scenarios:
        results: ScenarioResults = compute_scenario(bc, scenario)
        scenario.results = results.model_dump()
        flag_modified(scenario, "results")

    await db.flush()
    bc = await _get_bc_with_scenarios(db, use_case_id)
    return BusinessCaseRead.model_validate(bc)


# ─── Excel export ─────────────────────────────────────────────────────────────

def generate_excel(bc: BusinessCaseRead) -> bytes:
    """Generate an xlsx workbook from a fully-calculated BusinessCaseRead.

    Returns raw bytes suitable for a streaming HTTP response.
    Raises ValueError if no scenario has results.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for Excel export") from exc

    # Guard: at least one scenario must have been calculated
    scenarios_with_results = [s for s in bc.scenarios if s.results and s.results.get("monthly")]
    if not scenarios_with_results:
        raise ValueError("No scenarios have been calculated yet. Run Calculate first.")

    wb = Workbook()

    # ── Colour constants ──────────────────────────────────────────────────────
    HEADER_FILL = PatternFill("solid", fgColor="1A1D26")
    INPUT_FILL = PatternFill("solid", fgColor="4F7FFF")
    LABEL_FONT = Font(bold=True, color="F0F2F8")
    HEADER_FONT = Font(bold=True, color="F0F2F8")
    INPUT_FONT = Font(color="FFFFFF")
    CURRENCY_FMT = '$#,##0.00'
    PCT_FMT = '0.00%'

    def _header(ws: Any, row: int, col: int, value: str) -> None:
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left")

    def _label(ws: Any, row: int, col: int, value: str) -> None:
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = LABEL_FONT

    def _input_cell(ws: Any, row: int, col: int, value: Any) -> None:
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = INPUT_FILL
        cell.font = INPUT_FONT

    # ── Sheet 1: Assumptions ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Assumptions"
    ws1.column_dimensions["A"].width = 40
    ws1.column_dimensions["B"].width = 20

    rows: list[tuple[str, Any]] = [
        ("=== MODALITY PROFILE ===", None),
        ("Has Voice", bc.has_voice),
        ("Has Realtime Audio", bc.has_realtime_audio),
        ("Has Image Processing", bc.has_image_processing),
        ("Has Text Only", bc.has_text_only),
        ("STT Service", bc.stt_service or ""),
        ("TTS Service", bc.tts_service or ""),
        ("LLM Model", bc.llm_model or ""),
        ("IVR Service", bc.ivr_service or ""),
        ("", None),
        ("=== VOLUME & FTE ===", None),
        ("Weekly Volume (cases/calls)", bc.weekly_volume),
        ("Average Duration (minutes)", bc.avg_duration_minutes),
        ("FTE Count", bc.fte_count),
        ("Average FTE Annual Cost ($)", bc.avg_fte_annual_cost),
        ("FTE Monthly Overhead ($)", bc.fte_monthly_overhead),
        ("", None),
        ("=== TOKEN DENSITY ===", None),
        ("Token Density — Input (tokens/case)", bc.token_density_input),
        ("Token Density — Output (tokens/case)", bc.token_density_output),
        ("Caching Ratio", bc.caching_ratio),
        ("", None),
        ("=== COST MODEL ===", None),
        ("Implementation Cost ($)", bc.implementation_cost),
        ("Amortization Months", bc.implementation_amortization_months),
        ("Monthly Infrastructure Cost ($)", bc.monthly_infra_cost),
        ("Monthly Maintenance Cost ($)", bc.monthly_maintenance_cost),
        ("", None),
        ("=== GROWTH RATES ===", None),
        ("Volume Growth Rate YoY", bc.volume_growth_rate_yoy),
        ("Complexity Growth Rate YoY", bc.complexity_growth_rate_yoy),
        ("Inflation Rate YoY", bc.inflation_rate_yoy),
    ]

    for i, (label, value) in enumerate(rows, start=1):
        if label.startswith("==="):
            _header(ws1, i, 1, label)
        elif label == "":
            pass
        else:
            _label(ws1, i, 1, label)
            if value is not None:
                _input_cell(ws1, i, 2, value)

    # ── Sheet 2: Token Economics ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Token Economics")
    ws2.column_dimensions["A"].width = 38

    headers = ["Component"] + [s.name for s in bc.scenarios]
    for col, h in enumerate(headers, start=1):
        _header(ws2, 1, col, h)
        ws2.column_dimensions[get_column_letter(col)].width = 18

    price_rows = [
        ("LLM Input Price (per 1k tokens)", "llm_input_price_per_1k"),
        ("LLM Output Price (per 1k tokens)", "llm_output_price_per_1k"),
        ("Cached Input Price (per 1k tokens)", "cached_input_price_per_1k"),
        ("STT Price (per minute)", "stt_price_per_minute"),
        ("TTS Price (per 1k chars)", "tts_price_per_1k_chars"),
        ("IVR Price (per minute)", "ivr_price_per_minute"),
        ("Image Price (per image)", "image_price_per_image"),
    ]

    for row_i, (row_label, attr) in enumerate(price_rows, start=2):
        ws2.cell(row=row_i, column=1, value=row_label)
        for col_i, s in enumerate(bc.scenarios, start=2):
            ws2.cell(row=row_i, column=col_i, value=getattr(s, attr))

    # Cost per case row (from results)
    row_cpc = len(price_rows) + 2
    ws2.cell(row=row_cpc, column=1, value="Est. AI Cost per Case (Month 12)")
    for col_i, s in enumerate(bc.scenarios, start=2):
        monthly = (s.results or {}).get("monthly", [])
        if len(monthly) >= 12:
            cost = monthly[11].get("ai_cost_per_case", 0) if isinstance(monthly[11], dict) else getattr(monthly[11], "ai_cost_per_case", 0)
            cell = ws2.cell(row=row_cpc, column=col_i, value=cost)
            cell.number_format = CURRENCY_FMT

    # ── Sheet 3: Financial Model ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Financial Model")

    # Build header row
    fm_headers = ["Month", "Manual Labor Cost"]
    for s in bc.scenarios:
        fm_headers += [f"{s.name} — AI Total Cost", f"{s.name} — Monthly Savings"]

    for col, h in enumerate(fm_headers, start=1):
        _header(ws3, 1, col, h)
        ws3.column_dimensions[get_column_letter(col)].width = 22

    # Use first scenario's months as source for Manual Labor Cost
    first_monthly = []
    if bc.scenarios and bc.scenarios[0].results:
        first_monthly = bc.scenarios[0].results.get("monthly", [])

    for m in range(48):
        row = m + 2
        ws3.cell(row=row, column=1, value=m + 1)

        # Manual labor cost from first scenario
        if m < len(first_monthly):
            item = first_monthly[m]
            manual = item.get("manual_labor_cost", 0) if isinstance(item, dict) else getattr(item, "manual_labor_cost", 0)
        else:
            manual = 0.0
        cell_manual = ws3.cell(row=row, column=2, value=manual)
        cell_manual.number_format = CURRENCY_FMT

        col_offset = 3
        for s in bc.scenarios:
            monthly = (s.results or {}).get("monthly", [])
            if m < len(monthly):
                item = monthly[m]
                ai_cost = item.get("ai_total_cost", 0) if isinstance(item, dict) else getattr(item, "ai_total_cost", 0)
                savings = item.get("monthly_savings", 0) if isinstance(item, dict) else getattr(item, "monthly_savings", 0)
            else:
                ai_cost = 0.0
                savings = 0.0

            cell_ai = ws3.cell(row=row, column=col_offset, value=ai_cost)
            cell_ai.number_format = CURRENCY_FMT
            cell_sav = ws3.cell(row=row, column=col_offset + 1, value=savings)
            cell_sav.number_format = CURRENCY_FMT
            col_offset += 2

    # Totals row
    total_row = 50
    ws3.cell(row=total_row, column=1, value="TOTAL")
    ws3.cell(row=total_row, column=2).number_format = CURRENCY_FMT
    for col_i, s in enumerate(bc.scenarios):
        summary = s.results or {}
        total_inv = summary.get("total_ai_investment_48m", 0)
        total_sav = summary.get("total_savings_48m", 0)
        col_ai = 3 + col_i * 2
        col_sav = col_ai + 1
        cell_ai = ws3.cell(row=total_row, column=col_ai, value=total_inv)
        cell_ai.number_format = CURRENCY_FMT
        cell_sav = ws3.cell(row=total_row, column=col_sav, value=total_sav)
        cell_sav.number_format = CURRENCY_FMT

    # ── Sheet 4: FTE Impact ───────────────────────────────────────────────────
    ws4 = wb.create_sheet("FTE Impact")
    fte_headers = [
        "Month", "Coverage %", "AI Cases", "Human Cases",
        "Remaining FTEs", "Freed FTE Equivalent",
    ]
    for col, h in enumerate(fte_headers, start=1):
        _header(ws4, 1, col, h)
        ws4.column_dimensions[get_column_letter(col)].width = 22

    first_results_monthly = first_monthly
    for m in range(48):
        row = m + 2
        ws4.cell(row=row, column=1, value=m + 1)
        if m < len(first_results_monthly):
            item = first_results_monthly[m]
            if isinstance(item, dict):
                cov = item.get("coverage_pct", 0)
                ai = item.get("ai_cases", 0)
                human = item.get("human_cases", 0)
                rem_fte = item.get("remaining_fte", 0)
                freed = item.get("freed_capacity_fte", 0)
            else:
                cov = getattr(item, "coverage_pct", 0)
                ai = getattr(item, "ai_cases", 0)
                human = getattr(item, "human_cases", 0)
                rem_fte = getattr(item, "remaining_fte", 0)
                freed = getattr(item, "freed_capacity_fte", 0)
        else:
            cov = ai = human = rem_fte = freed = 0

        ws4.cell(row=row, column=2, value=cov).number_format = PCT_FMT
        ws4.cell(row=row, column=3, value=ai).number_format = '#,##0.00'
        ws4.cell(row=row, column=4, value=human).number_format = '#,##0.00'
        ws4.cell(row=row, column=5, value=rem_fte).number_format = '#,##0.00'
        ws4.cell(row=row, column=6, value=freed).number_format = '#,##0.00'

    # ── Sheet 5: ROI Summary ──────────────────────────────────────────────────
    ws5 = wb.create_sheet("ROI Summary")
    summary_headers = ["Metric"] + [s.name for s in bc.scenarios]
    for col, h in enumerate(summary_headers, start=1):
        _header(ws5, 1, col, h)
        ws5.column_dimensions[get_column_letter(col)].width = 25

    summary_rows = [
        ("Break-even Month", "break_even_month"),
        ("ROI at 12 Months (%)", "roi_12m"),
        ("ROI at 24 Months (%)", "roi_24m"),
        ("ROI at 36 Months (%)", "roi_36m"),
        ("ROI at 48 Months (%)", "roi_48m"),
        ("Total Savings 48M ($)", "total_savings_48m"),
        ("Cost Savings % (vs Manual)", "cost_savings_pct"),
        ("Total AI Investment 48M ($)", "total_ai_investment_48m"),
        ("Total Manual Cost 48M ($)", "total_manual_cost_48m"),
        ("Remaining FTEs", "remaining_fte"),
        ("Freed FTE Equivalent", "freed_capacity_fte"),
    ]

    for row_i, (metric_label, key) in enumerate(summary_rows, start=2):
        ws5.cell(row=row_i, column=1, value=metric_label)
        for col_i, s in enumerate(bc.scenarios, start=2):
            val = (s.results or {}).get(key)
            cell = ws5.cell(row=row_i, column=col_i, value=val)
            if key in ("total_savings_48m", "total_ai_investment_48m", "total_manual_cost_48m"):
                cell.number_format = CURRENCY_FMT
            elif key in ("roi_12m", "roi_24m", "roi_36m", "roi_48m", "cost_savings_pct"):
                cell.number_format = '#,##0.00"%"'

    # ── Serialise to bytes ────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
